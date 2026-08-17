"""What the visit form must keep true, in the schema and in the rendered workbook.

The workbook is the only place several of these errors are cheap to catch, so the
tests assert the *refusals* — an unsigned coordinate, a date Excel can reparse, a
camera unit named like a station — not merely that the file was produced.
"""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from camtrap import stations, visit_schema
from camtrap.visit_schema import (
    LAT_BOUNDS,
    LON_BOUNDS,
    REQ_ALWAYS,
    STATION_FIELDS,
    VISIT_FIELDS,
    by_label,
    read_coordinate,
)
from setup.build_visit_template import build

ALL_FIELDS = VISIT_FIELDS + STATION_FIELDS


class TestSchema(unittest.TestCase):

    def test_labels_are_unique_within_each_sheet(self):
        for fields in (VISIT_FIELDS, STATION_FIELDS):
            labels = [f.label for f in fields]
            self.assertEqual(len(labels), len(set(labels)))

    def test_columns_are_unique_within_each_sheet(self):
        for fields in (VISIT_FIELDS, STATION_FIELDS):
            cols = [f.column for f in fields]
            self.assertEqual(len(cols), len(set(cols)))

    def test_at_most_one_validation_rule_per_field(self):
        """Excel honours one validation per cell; two declared rules means one is
        silently dropped, and it would be the one written second."""
        for f in ALL_FIELDS:
            declared = sum(bool(x) for x in (f.options, f.bounds, f.length, f.prefix))
            self.assertLessEqual(declared, 1, f'{f.column} declares {declared} rules')

    def test_the_closed_campaign_is_never_asked_for(self):
        """It is the campaign the previous visit to that station opened, so asking
        for it would let the sheet contradict itself."""
        cols = {f.column for f in VISIT_FIELDS}
        self.assertIn('campaign_opened', cols)
        self.assertNotIn('campaign_closed', cols)

    def test_visit_rows_never_ask_for_standing_site_facts(self):
        """grid_id and elevation belong to the site, live once in `Estaciones`, and
        must not be retyped into every visit."""
        visit_cols = {f.column for f in VISIT_FIELDS}
        for column in ('grid_id', 'elevation_m'):
            self.assertNotIn(column, visit_cols)
            self.assertIn(column, {f.column for f in STATION_FIELDS})

    def test_no_conclusion_columns(self):
        """The whole point of the form: raw readings, never a verdict."""
        cols = {f.column for f in VISIT_FIELDS}
        for banned in ('clock_state', 'clock_action', 'clock_offset_hours'):
            self.assertNotIn(banned, cols)

    def test_both_clocks_are_obligatory(self):
        for col in ('visit_date', 'visit_time', 'camera_datetime_observed'):
            f = next(x for x in VISIT_FIELDS if x.column == col)
            self.assertEqual(f.required, REQ_ALWAYS, col)

    def test_coordinates_are_bounded_to_the_southern_hemisphere(self):
        lat = next(f for f in VISIT_FIELDS if f.column == 'lat')
        lon = next(f for f in VISIT_FIELDS if f.column == 'lon')
        self.assertLess(lat.bounds[1], 0, 'a positive latitude must be refusable')
        self.assertLess(lon.bounds[1], 0, 'a positive longitude must be refusable')
        # Bosque Pehuén itself must be admissible.
        self.assertTrue(lat.bounds[0] <= -39.4417 <= lat.bounds[1])
        self.assertTrue(lon.bounds[0] <= -71.7420 <= lon.bounds[1])

    def test_station_ids_come_from_the_naming_convention(self):
        self.assertIn(stations.canonical_id(1), visit_schema.STATION_IDS)
        self.assertIn(stations.canonical_id(27), visit_schema.STATION_IDS)
        for sid in visit_schema.STATION_IDS:
            self.assertTrue(stations.is_canonical(sid), sid)

    def test_unit_id_prefix_cannot_be_mistaken_for_a_station(self):
        f = next(x for x in VISIT_FIELDS if x.column == 'camera_unit_id')
        self.assertTrue(f.prefix)
        self.assertFalse(stations.is_canonical(f.prefix + '18'))

    def test_every_field_explains_itself(self):
        for f in ALL_FIELDS:
            self.assertTrue(f.why.strip(), f'{f.column} has no glossary text')

    def test_by_label_round_trips(self):
        for f in VISIT_FIELDS:
            self.assertIs(by_label(f.label), f)

    def test_by_label_names_an_unknown_header(self):
        with self.assertRaises(KeyError) as ctx:
            by_label('Grilla')
        self.assertIn('Grilla', str(ctx.exception))


class TestCoordinateRule(unittest.TestCase):
    """CT26 sat 19 km outside the reserve for a year because 39°25'44.7" was typed
    as 39.25447. Repaired in the platform 2026-04-15; `build_field_notes.py` was
    written in August and re-sourced the same cell, which is the failure this rule
    exists to make impossible."""

    def test_ct26_is_recognised_as_dms_and_converted(self):
        lat, lat_flag = read_coordinate(39.25447, 'lat')
        lon, lon_flag = read_coordinate(71.44562, 'lon')
        self.assertAlmostEqual(lat, -39.42908, places=5)
        self.assertAlmostEqual(lon, -71.74894, places=5)
        for flag in (lat_flag, lon_flag):
            self.assertIn('coord_dms_as_decimal', flag)

    def test_a_plain_decimal_is_left_alone_and_signed(self):
        value, flag = read_coordinate(39.45183, 'lat')
        self.assertAlmostEqual(value, -39.45183, places=5)
        self.assertEqual(flag, '')

    def test_an_already_signed_value_is_idempotent(self):
        self.assertEqual(read_coordinate(-39.45183, 'lat'),
                         read_coordinate(39.45183, 'lat'))

    def test_a_seconds_field_over_sixty_is_not_a_dms_reading(self):
        """20 of the 52 historical coordinates have digits that cannot be minutes
        and seconds (39.43796 -> 43' 79.6"). Treating "no DMS reading" as "the DMS
        reading agrees" flagged every one of them ambiguous."""
        value, flag = read_coordinate(39.43796, 'lat')
        self.assertAlmostEqual(value, -39.43796, places=5)
        self.assertEqual(flag, '')

    def test_a_value_outside_the_reserve_is_refused_not_guessed(self):
        value, flag = read_coordinate(12.5, 'lat')
        self.assertIsNone(value)
        self.assertIn('coord_out_of_range', flag)

    def test_blank_is_not_an_error(self):
        for blank in (None, '', '   '):
            self.assertEqual(read_coordinate(blank, 'lat'), (None, ''))

    def test_unparseable_text_is_named(self):
        value, flag = read_coordinate('sin dato', 'lat')
        self.assertIsNone(value)
        self.assertIn('coord_unparseable', flag)

    def test_bounds_are_tight_enough_for_the_rule_to_decide(self):
        """A country-wide box makes both readings of 39.25447 plausible and the DMS
        test can decide nothing. This is the invariant that widening the bounds
        would silently break."""
        for bounds, kind in ((LAT_BOUNDS, 'lat'), (LON_BOUNDS, 'lon')):
            self.assertLess(bounds[1] - bounds[0], 1.0, f'{kind} box too wide')

    def test_the_whole_historical_record_resolves(self):
        import csv
        path = Path(__file__).resolve().parents[1] / 'data/campaigns/field_notes.csv'
        seen = 0
        with open(path, encoding='utf-8', newline='') as fh:
            for row in csv.DictReader(fh):
                for column, kind in (('lat', 'lat'), ('lon', 'lon')):
                    if not (row[column] or '').strip():
                        continue
                    seen += 1
                    value, _ = read_coordinate(row[column], kind)
                    self.assertIsNotNone(value, f'{row["station_id"]} {column}')
                    lo, hi = LAT_BOUNDS if kind == 'lat' else LON_BOUNDS
                    self.assertTrue(lo <= value <= hi)
        self.assertGreater(seen, 0, 'fixture found no coordinates to check')


class TestWorkbook(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        from openpyxl import load_workbook
        cls._tmp = tempfile.TemporaryDirectory()
        path = build(Path(cls._tmp.name) / 'registro.xlsx', rows=5)
        cls.wb = load_workbook(path)
        cls.visits = cls.wb['Visitas']
        cls.stations_sheet = cls.wb['Estaciones']

    @classmethod
    def tearDownClass(cls):
        cls._tmp.cleanup()

    def _column_of(self, column: str) -> str:
        from openpyxl.utils import get_column_letter
        idx = next(i for i, f in enumerate(VISIT_FIELDS, start=1) if f.column == column)
        return get_column_letter(idx)

    def test_headers_match_the_schema_in_order(self):
        got = [self.visits.cell(1, i).value for i in range(1, len(VISIT_FIELDS) + 1)]
        self.assertEqual(got, [f.label for f in VISIT_FIELDS])

    def test_date_cells_are_text_so_excel_cannot_reparse_them(self):
        """The legacy workbook's dangerous cells were the ones Excel had already
        parsed with the machine locale — a wrong reading that looks clean."""
        for f in VISIT_FIELDS:
            if not f.is_text:
                continue
            letter = self._column_of(f.column)
            self.assertEqual(self.visits[f'{letter}2'].number_format, '@', f.column)

    def test_every_declared_rule_reaches_the_workbook(self):
        rules = {str(dv.sqref): dv for dv in self.visits.data_validations.dataValidation}
        for f in VISIT_FIELDS:
            if not any((f.options, f.bounds, f.length, f.prefix)):
                continue
            letter = self._column_of(f.column)
            self.assertIn(f'{letter}2:{letter}6', rules, f'{f.column} lost its rule')

    def test_a_positive_latitude_is_out_of_range(self):
        dv = next(d for d in self.visits.data_validations.dataValidation
                  if str(d.sqref).startswith(self._column_of('lat') + '2:'))
        self.assertEqual(dv.operator, 'between')
        self.assertLess(float(dv.formula2), 0)

    def test_hidden_list_sheet_backs_every_dropdown(self):
        self.assertEqual(self.wb['Listas'].sheet_state, 'hidden')
        for f in VISIT_FIELDS:
            if not f.has_list:
                continue
            letter = self._column_of(f.column)
            dv = next(d for d in self.visits.data_validations.dataValidation
                      if str(d.sqref) == f'{letter}2:{letter}6')
            self.assertIn('Listas', dv.formula1)

    def test_example_rows_are_on_their_own_sheet(self):
        """An example row inside `Visitas` would be ingested as a real visit."""
        self.assertIn('Ejemplo', self.wb.sheetnames)
        for row in self.visits.iter_rows(min_row=2, values_only=True):
            self.assertTrue(all(v is None for v in row))

    def test_glossary_documents_every_field(self):
        glossary = self.wb['Glosario']
        first = {glossary.cell(r, 1).value for r in range(1, glossary.max_row + 1)}
        for f in ALL_FIELDS:
            self.assertIn(f.label, first, f'{f.column} missing from the glossary')

    def test_there_is_exactly_one_sheet_to_fill(self):
        """Every other visible sheet is reference. A second fillable sheet would
        reintroduce the per-salida copy this design exists to remove."""
        visible = [s.title for s in self.wb.worksheets if s.sheet_state == 'visible']
        self.assertEqual(visible, ['Visitas', 'Ejemplo', 'Estaciones', 'Glosario'])

    def test_the_station_sheet_is_prefilled_from_the_registry(self):
        registry = stations.registry()
        self.assertTrue(registry, 'estaciones.csv is missing or empty')
        got = {self.stations_sheet.cell(r, 1).value
               for r in range(2, self.stations_sheet.max_row + 1)}
        self.assertEqual(got, set(registry))

    def test_ct26_carries_the_repaired_coordinate(self):
        row = next(r for r in range(2, self.stations_sheet.max_row + 1)
                   if self.stations_sheet.cell(r, 1).value == 'CT26')
        lat_col = next(i for i, f in enumerate(STATION_FIELDS, 1) if f.column == 'lat')
        lon_col = next(i for i, f in enumerate(STATION_FIELDS, 1) if f.column == 'lon')
        self.assertAlmostEqual(self.stations_sheet.cell(row, lat_col).value,
                               -39.42908, places=5)
        self.assertAlmostEqual(self.stations_sheet.cell(row, lon_col).value,
                               -71.74894, places=5)


if __name__ == '__main__':
    unittest.main()
