"""
setup/build_visit_template.py — render the visit schema to the fillable workbook.

WHAT THIS OWNS

    How a `camtrap.visit_schema` field becomes Excel: which validation rule enforces
    it, whether the cell must resist Excel's date parser, how the glossary reads.
    It knows openpyxl; it does not know the field list. Adding a column is an edit
    to `visit_schema.py` and a re-run of this script.

ONE FILE, ONE SHEET TO FILL, NO COPIES

    The output is a single workbook, `Registro de visitas CT.xlsx`, named as the
    successor to the legacy `Registro de monitoreo CT.xlsx`. Rows accumulate on the
    `Visitas` sheet forever — a salida is not a new file, a new sheet, or a new
    folder, so the instruction to a field coordinator is one sentence: *abre el
    archivo y agrega una fila por cámara visitada*.

    This is a deliberate reversal. Per-salida copies would mean a naming convention,
    a folder tree and a README nobody reads, and the predictable outcome is someone
    duplicating a sheet by hand anyway. Accumulating rows removes the step instead of
    policing it — and if a sheet does get duplicated, every sheet whose headers match
    is still readable, so the habit stops being a failure mode.

    `campaign_closed` is not asked at all: the campaign a visit closes is the one the
    previous visit to that same station opened, so it is derived and can never
    contradict the record.

WHY GENERATED AND NOT DRAWN BY HAND

    A hand-authored .xlsx is a binary blob: it cannot be reviewed in a diff, and its
    dropdowns and its glossary drift apart the first time someone edits one and not
    the other. Generating both from one declaration makes that drift impossible.

WHAT THE WORKBOOK REFUSES, AND WHY THAT IS THE POINT

    Every rule here exists because the corresponding error is already in the record:

      unsigned coordinates   the historical CSV held lat 39.45 / lon 71.72 — China.
      DMS typed as decimal   CT26 sat 19 km outside the reserve for a year because
                             39°25'44.7" was copied as 39.25447.
      locale-parsed dates    the legacy workbook's most dangerous cells were the ones
                             Excel had already turned into datetimes — a wrong
                             reading that looks clean. Date cells are TEXT.
      unit/station collision May 2026: station CT23 received camera unit 18, station
                             CT18 received unit 28.
      a missing visit hour   27 of 27 otoño 2026 opening visits. Now obligatory.

    Refusing at the tree costs one retype. Refusing at ingest costs a season.

Usage:  python setup/build_visit_template.py [--out PATH] [--rows N]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from camtrap import stations  # noqa: E402
from camtrap.visit_schema import (  # noqa: E402
    REQ_ALWAYS,
    SCHEMA_VERSION,
    STATION_FIELDS,
    VISIT_FIELDS,
    VisitField,
)

OUT_XLSX = Path('data/campaigns/Registro de visitas CT.xlsx')
DATA_ROWS = 60

SH_VISITS   = 'Visitas'
SH_EXAMPLE  = 'Ejemplo'
SH_STATIONS = 'Estaciones'
SH_GLOSSARY = 'Glosario'
SH_LISTS    = 'Listas'

INK        = '1F3B2E'   # header background, obligatory
INK_COND   = '7A5C18'   # header background, conditional or optional
PAPER      = 'F2EFE6'
ALERT      = 'FFC7CE'   # conditional-formatting highlight
RULE_LINE  = 'C9C2B2'

_thin = Side(style='thin', color=RULE_LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _clip(text: str, limit: int) -> str:
    """Excel silently corrupts a validation prompt over its limit, so clip first."""
    return text if len(text) <= limit else text[: limit - 1] + '…'


def _header_style(ws, col: int, f: VisitField) -> None:
    cell = ws.cell(row=1, column=col, value=f.label)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', fgColor=INK if f.required == REQ_ALWAYS else INK_COND)
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = f.width


def _number_format(f: VisitField) -> str:
    if f.is_text:
        return '@'
    if f.column in ('lat', 'lon'):
        return '0.00000'
    if f.column in ('bearing_deg', 'elevation_m', 'grid_id'):
        return '0'
    if f.bounds:
        return '0.0'
    return 'General' if f.has_list else '@'


def _validation(f: VisitField, ref: str, lists: dict[str, str]) -> DataValidation | None:
    """One rule per column — Excel honours only one validation per cell.

    Precedence is by construction, not by choice: a field declares exactly one of
    options / bounds / length / prefix.
    """
    common = dict(
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
        promptTitle=_clip(f.label, 32),
        prompt=_clip(f'{f.fmt}\nObligatorio: {f.required}\nEj.: {f.example or "—"}', 255),
        errorTitle=_clip(f.label, 32),
    )

    if f.has_list:
        dv = DataValidation(
            type='list', formula1=lists[f.column],
            error=_clip(f'Elegir una opción de la lista: {", ".join(f.options)}', 255),
            **common)
    elif f.bounds:
        lo, hi = f.bounds
        kind = 'whole' if f.column == 'bearing_deg' else 'decimal'
        dv = DataValidation(
            type=kind, operator='between', formula1=lo, formula2=hi,
            error=_clip(f'Debe ser un número entre {lo} y {hi}. {f.fmt}.', 255),
            **common)
    elif f.length:
        dv = DataValidation(
            type='textLength', operator='equal', formula1=f.length,
            error=_clip(f'Formato exacto {f.fmt} — {f.length} caracteres. '
                        f'Ej.: {f.example}', 255),
            **common)
    elif f.prefix:
        col = ref.split('2:')[0]
        dv = DataValidation(
            type='custom',
            formula1=f'=OR(ISBLANK({col}2),LEFT({col}2,{len(f.prefix)})="{f.prefix}")',
            error=_clip(f'Debe empezar con "{f.prefix}" — es el ID del EQUIPO, no de '
                        f'la estación. Ej.: {f.example}', 255),
            **common)
    else:
        return None

    dv.add(ref)
    return dv


def _build_lists(wb: Workbook) -> dict[str, str]:
    """Write every option list to a hidden sheet, return its absolute range."""
    ws = wb.create_sheet(SH_LISTS)
    ws['A1'] = 'Generado por setup/build_visit_template.py — no editar.'
    refs: dict[str, str] = {}
    col = 1
    for f in VISIT_FIELDS + STATION_FIELDS:
        if not f.has_list or f.column in refs:
            continue
        letter = get_column_letter(col)
        ws.cell(row=2, column=col, value=f.column).font = Font(bold=True)
        for i, opt in enumerate(f.options, start=3):
            ws.cell(row=i, column=col, value=opt)
        refs[f.column] = f"'{SH_LISTS}'!${letter}$3:${letter}${2 + len(f.options)}"
        col += 1
    ws.sheet_state = 'hidden'
    return refs


def _build_grid(wb: Workbook, title: str, fields: tuple[VisitField, ...],
                lists: dict[str, str], rows: int, values: list[list] | None = None):
    """A header row, validated blank rows beneath it, optionally pre-filled."""
    ws = wb.create_sheet(title)
    last_row = rows + 1

    for i, f in enumerate(fields, start=1):
        _header_style(ws, i, f)
        letter = get_column_letter(i)
        ref = f'{letter}2:{letter}{last_row}'
        fmt = _number_format(f)

        for r in range(2, last_row + 1):
            cell = ws.cell(row=r, column=i)
            cell.number_format = fmt
            cell.border = BORDER
            cell.alignment = Alignment(
                wrap_text=f.column == 'notes', vertical='center',
                horizontal='left' if f.column == 'notes' else 'center')
            if r % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=PAPER)

        dv = _validation(f, ref, lists)
        if dv is not None:
            ws.add_data_validation(dv)

        # A text-formatted cell that Excel nonetheless stored as a number means the
        # value was pasted and reparsed — the exact failure that made the legacy
        # workbook's clean-looking dates untrustworthy. Flag it where it happens.
        if f.is_text:
            ws.conditional_formatting.add(ref, FormulaRule(
                formula=[f'AND(NOT(ISBLANK({letter}2)),ISNUMBER({letter}2))'],
                fill=PatternFill('solid', fgColor=ALERT), stopIfTrue=False))

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 46
    ws.auto_filter.ref = f'A1:{get_column_letter(len(fields))}{last_row}'

    for r, row_values in enumerate(values or [], start=2):
        for c, value in enumerate(row_values, start=1):
            ws.cell(row=r, column=c, value=value)

    return ws


def _station_rows() -> list[list]:
    """The registry, rendered as `Estaciones` rows. Numbers stay numbers so the
    sheet sorts and filters; blanks stay blank because they mean NOT RECORDED."""
    numeric = {'lat', 'lon', 'height_m', 'detection_distance_m'}
    integer = {'grid_id', 'elevation_m', 'bearing_deg'}
    rows = []
    for sid, entry in sorted(stations.registry().items()):
        row = []
        for f in STATION_FIELDS:
            raw = (entry.get(f.column) or '').strip()
            if not raw:
                row.append(None)
            elif f.column in integer:
                row.append(int(float(raw)))
            elif f.column in numeric:
                row.append(float(raw))
            else:
                row.append(raw)
        rows.append(row)
    return rows


EXAMPLE_ROWS = [
    # A normal revision with a clock that is NOT adjusted — the reading is one hour
    # behind the reference and stays that way on the sheet. This is the case the old
    # form destroyed by asking for `clock_action=shifted, -1.0` instead.
    ['CT01', '2026-05-15', '09:40', 'revision', 'primavera_2026', 'TA, SC', 'CAM-01',
     'si', '2026-05-15 08:40', 'no', '', 'si', 'si', 'no',
     None, None, None, None, None,
     'Sin novedad. La pantalla va 1 h atrás del teléfono; NO se ajustó.'],
    # A dead camera: no screen to read, so the datetime is blank and `camera_working`
    # carries the reason. The unit that replaced it is named with its CAM- prefix.
    ['CT18', '2026-05-15', '12:10', 'revision', 'primavera_2026', 'TA, SC', 'CAM-28',
     'no', '', 'no', '', 'si', 'si', 'no',
     None, None, None, None, None,
     'No prendía, se cambió la CT a otra. Equipo retirado: CAM-18.'],
    # A move: the five position columns are filled only here, and the coordinates
    # carry their signs.
    ['CT07', '2026-05-16', '11:05', 'revision', 'primavera_2026', 'TA', 'CAM-07',
     'si', '2026-05-16 11:05', 'no', '', 'si', 'no', 'si',
     -39.44921, -71.73588, 1.6, 210, 4.5,
     'Se corrió 30 m al sur del sendero, la vegetación tapaba el lente.'],
]


def _build_glossary(wb: Workbook):
    ws = wb.create_sheet(SH_GLOSSARY)
    for i, w in enumerate((34, 24, 30, 26, 30, 20, 96), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws['A1'] = 'Glosario — Registro de visitas a cámaras trampa'
    ws['A1'].font = Font(bold=True, size=14, color=INK)
    ws['A2'] = (
        f'Versión {SCHEMA_VERSION}.  '
        'CÓMO SE USA: en la hoja «Visitas», agregar una fila por cada cámara '
        'visitada. No se crea un archivo nuevo ni una hoja nueva por salida — las '
        'filas se van acumulando siempre en la misma hoja.  '
        'La hoja «Estaciones» es de consulta: ahí están la grilla y las coordenadas '
        'de cada sitio, no hay que copiarlas a la visita.  '
        'Encabezado VERDE = obligatorio siempre.  Encabezado CAFÉ = obligatorio sólo '
        'en el caso indicado, o bien opcional.  '
        'Las fechas se escriben como texto ISO (AAAA-MM-DD) para que Excel no las '
        'reinterprete según el idioma del computador; una celda que se pinte de rojo '
        'es una fecha que Excel convirtió y hay que volver a escribir.')
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 92

    head = ('Campo (lo que dice la planilla)', 'Columna destino', 'Formato',
            'Opciones válidas', 'Obligatorio', 'Ejemplo', 'Por qué se pide')

    def section(title: str, fields: tuple[VisitField, ...], row: int) -> int:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=11, color=INK)
        row += 1
        for c, text in enumerate(head, start=1):
            h = ws.cell(row=row, column=c, value=text)
            h.font = Font(bold=True, color='FFFFFF', size=10)
            h.fill = PatternFill('solid', fgColor=INK)
            h.alignment = Alignment(wrap_text=True, vertical='center')
            h.border = BORDER
        row += 1
        for f in fields:
            values = (f.label, f.column, f.fmt, ' / '.join(f.options) or '—',
                      f.required, f.example or '—', f.why)
            for c, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = BORDER
                cell.font = Font(size=9, bold=c == 1)
                if c == 1:
                    cell.fill = PatternFill(
                        'solid',
                        fgColor=PAPER if f.required == REQ_ALWAYS else 'FBF3DE')
            ws.row_dimensions[row].height = max(30, 12 * (len(f.why) // 90 + 1) + 18)
            row += 1
        return row + 2

    row = section(f'Hoja «{SH_VISITS}» — una fila por estación y por visita',
                  VISIT_FIELDS, 4)
    section(f'Hoja «{SH_STATIONS}» — datos fijos del sitio, sólo de consulta',
            STATION_FIELDS, row)

    ws.freeze_panes = 'A5'
    return ws


def build(out_path: Path = OUT_XLSX, rows: int = DATA_ROWS) -> Path:
    """Write the fillable workbook. Overwrites; the file is a build artefact."""
    wb = Workbook()
    wb.remove(wb.active)

    lists = _build_lists(wb)
    registry = _station_rows()
    _build_grid(wb, SH_VISITS, VISIT_FIELDS, lists, rows)
    _build_grid(wb, SH_EXAMPLE, VISIT_FIELDS, lists, len(EXAMPLE_ROWS), EXAMPLE_ROWS)
    _build_grid(wb, SH_STATIONS, STATION_FIELDS, lists, len(registry), registry)
    _build_glossary(wb)

    # `Listas` is built first because every validation references it, but the person
    # opening the file must land on the sheet they fill.
    wb.move_sheet(SH_LISTS, offset=len(wb.sheetnames) - 1)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument('--out', type=Path, default=OUT_XLSX)
    ap.add_argument('--rows', type=int, default=DATA_ROWS,
                    help='filas en blanco en la hoja Visitas')
    args = ap.parse_args()

    path = build(args.out, args.rows)
    obligatory = sum(1 for f in VISIT_FIELDS if f.required == REQ_ALWAYS)
    print(f'{path}: v{SCHEMA_VERSION}, {len(VISIT_FIELDS)} columnas de visita '
          f'({obligatory} obligatorias), {len(stations.registry())} estaciones '
          f'en la hoja de consulta, {args.rows} filas en blanco.')


if __name__ == '__main__':
    main()
