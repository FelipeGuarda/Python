"""
camtrap/visit_form.py — how a filled visit workbook becomes field-record rows.

WHAT THIS OWNS

    The translation between the sheet a technician fills in the field and the rows
    `field_notes.csv` holds: resolving Spanish headers back to columns, enforcing the
    form's own conditional obligations, normalising what Excel did to a date, and
    deciding when a workbook is refused outright.

    It does NOT own the field list, the wording, the allowed answers or the
    requirement rules — `camtrap/visit_schema.py` declares all four and this module
    asks. Adding a covariate is an edit there and changes nothing here, which only
    holds because no form column is named literally below.

ALL OR NOTHING, AND EVERY PROBLEM AT ONCE

    A workbook is validated whole and either every row lands or none does. Half a
    salida in the record is worse than none: the missing half is invisible, while a
    refused file is a message. `VisitFormError` carries every problem found rather
    than the first, because the person fixing it is at a desk with the sheet in front
    of them and wants the list.

WHY APPENDING NEVER REWRITES

    The 107 legacy rows carry curation that exists nowhere else — CT27's retrieval
    date was deduced from the order of the other cameras' last frames and lives in no
    field sheet at all. So this module only ever appends, never rewrites or reorders,
    and refuses a visit already in the file. A future salida cannot reach the history.

WHAT `source_sheet` AND `data_flags` ARE FOR

    Neither is collected on the form; both are written here. `source_sheet` names the
    workbook a row came from, so a row can always be traced to the paper it was
    transcribed from, and `data_flags` carries what could be read but not settled —
    an out-of-range coordinate, a date Excel had already reparsed. `anchors.Visit`
    reads both, which is why they survive the reshape.
"""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime, time
from pathlib import Path

from camtrap import visit_schema
from camtrap.anchors import FIELD_NOTES_FILENAME, RECORDED_CLOSE_COLUMN

#: The sheet the technician fills. `setup/build_visit_template.py` names it.
SHEET = 'Visitas'

#: Written by this module, not collected on the form. See the module docstring.
PROVENANCE_COLUMNS = ('source_sheet', 'data_flags')

#: The shape of `field_notes.csv`: the form, in form order, plus the provenance two.
FIELD_NOTES_COLUMNS = (
    tuple(f.column for f in visit_schema.VISIT_FIELDS) + PROVENANCE_COLUMNS
)

#: A visit is one station on one date at one time. Two rows sharing these three are
#: the same visit transcribed twice, which is what re-ingesting a workbook does.
VISIT_KEY = ('station_id', 'visit_date', 'visit_time')

_DATE_OUT = '%Y-%m-%d'
_TIME_OUT = '%H:%M'
_DATETIME_OUT = '%Y-%m-%d %H:%M'

_BOOLEAN_COLUMNS = tuple(
    f.column for f in visit_schema.VISIT_FIELDS
    if f.options in (visit_schema.SI_NO, visit_schema.SI_NO_NS)
)


class VisitFormError(Exception):
    """One workbook's problems, all of them."""

    def __init__(self, path: Path, problems: list[str]):
        self.path = path
        self.problems = list(problems)
        listed = '\n  - '.join(self.problems)
        super().__init__(f'{path.name}: {len(self.problems)} problema(s)\n  - {listed}')


# =============================================================================
# Reading one cell
# =============================================================================

def _blank(raw) -> bool:
    return raw is None or (isinstance(raw, str) and not raw.strip())


def _text(raw) -> str:
    """A cell as the string the CSV should hold.

    Excel silently retypes what looks like a number, so `'2026-05-15'` can arrive as
    a `datetime` and `'09:40'` as a `time`. Formatting them back is not cosmetic:
    `str(datetime)` would write `2026-05-15 00:00:00` into a date column.
    """
    if _blank(raw):
        return ''
    if isinstance(raw, datetime):
        return raw.strftime(_DATETIME_OUT if (raw.hour or raw.minute)
                            else _DATE_OUT)
    if isinstance(raw, date):
        return raw.strftime(_DATE_OUT)
    if isinstance(raw, time):
        return raw.strftime(_TIME_OUT)
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def _read_date(raw, column: str, where: str, problems: list[str]) -> str:
    value = _text(raw)
    if not value:
        return ''
    try:
        return datetime.strptime(value, _DATE_OUT).strftime(_DATE_OUT)
    except ValueError:
        problems.append(f'{where}: {column} = {value!r} no es {visit_schema.FMT_DATE}')
        return ''


def _read_time(raw, column: str, where: str, problems: list[str]) -> str:
    value = _text(raw)
    if not value:
        return ''
    for fmt in (_TIME_OUT, '%H:%M:%S'):
        try:
            return datetime.strptime(value, fmt).strftime(_TIME_OUT)
        except ValueError:
            pass
    problems.append(f'{where}: {column} = {value!r} no es {visit_schema.FMT_TIME}')
    return ''


def _read_datetime(raw, column: str, where: str, problems: list[str]) -> str:
    """A raw clock reading. Absurd values are kept: CT18's screen said 2017."""
    value = _text(raw)
    if not value:
        return ''
    for fmt in (_DATETIME_OUT, '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(value, fmt).strftime(_DATETIME_OUT)
        except ValueError:
            pass
    problems.append(
        f'{where}: {column} = {value!r} no es {visit_schema.FMT_DATETIME}')
    return ''


def _read_number(raw, field, where: str, problems: list[str]) -> str:
    value = _text(raw)
    if not value:
        return ''
    try:
        number = float(value)
    except ValueError:
        problems.append(f'{where}: {field.column} = {value!r} no es un número')
        return ''
    lo, hi = field.bounds
    if not lo <= number <= hi:
        problems.append(f'{where}: {field.column} = {number} fuera de rango [{lo}, {hi}]')
        return ''
    return str(int(number)) if float(number).is_integer() else str(number)


def _read_choice(raw, field, where: str, problems: list[str]) -> str:
    """A list cell, in the vocabulary's own spelling, translated where the CSV has one.

    Matching is case-insensitive and returns the declared option rather than what was
    typed, so `ct01` becomes `CT01`. The list is the authority on spelling; a
    dropdown that was overtyped in the field should not create a second dialect.
    """
    value = _text(raw)
    if not value:
        return ''
    match = next((o for o in field.options if str(o).lower() == value.lower()), None)
    if match is None:
        problems.append(
            f'{where}: {field.column} = {value!r} no está en la lista '
            f'({", ".join(str(o) for o in field.options)})')
        return ''
    return visit_schema.CSV_BOOLEAN.get(match, match) \
        if field.column in _BOOLEAN_COLUMNS else match


# =============================================================================
# Reading one row
# =============================================================================

def _requires_placement(row: dict) -> bool:
    """`REQ_IF_MOVED`, as the form states it: moved, or a new installation."""
    return row.get('moved') == 'yes' or row.get('visit_type') == 'instalacion'


def _check_obligations(row: dict, where: str, problems: list[str]) -> None:
    """The form's four requirement classes, applied to one finished row.

    Two refinements the requirement column cannot express on its own, both taken
    from the fields' own `why` text rather than invented here:

      * `camera_datetime_observed` is blank when the camera would not switch on —
        the form tells the technician to leave it blank and answer `camera_working`
        = no, so demanding it would demand a reading that does not exist.
      * `campaign_opened` must be BLANK on a `retiro`: the vocabulary defines that
        visit as leaving the site without equipment, so a campaign named there
        would contradict the visit type. The requirement column says `siempre`
        because every other visit type does leave a card in the ground.
    """
    for field in visit_schema.VISIT_FIELDS:
        value = row.get(field.column, '')
        if field.required == visit_schema.REQ_ALWAYS:
            if field.column == 'camera_datetime_observed' and row.get('camera_working') == 'no':
                continue
            if field.column == 'campaign_opened' and row.get('visit_type') == 'retiro':
                continue
            if not value:
                problems.append(f'{where}: falta {field.column} (obligatorio)')
        elif field.required == visit_schema.REQ_IF_ADJUSTED:
            if row.get('clock_adjusted') == 'yes' and not value:
                problems.append(
                    f'{where}: falta {field.column}; el reloj se ajustó, así que sin '
                    'esta lectura el desfase queda desconocido desde ese instante')
        elif field.required == visit_schema.REQ_IF_MOVED:
            if _requires_placement(row) and not value:
                problems.append(f'{where}: falta {field.column} (se movió o es instalación)')

    if row.get('visit_type') == 'retiro' and row.get('campaign_opened'):
        problems.append(
            f'{where}: un retiro deja el sitio sin equipo, así que campaign_opened '
            f'debe quedar vacío (dice {row["campaign_opened"]!r})')


def _read_row(values: dict, where: str, source_sheet: str,
              problems: list[str]) -> dict:
    """One sheet row as a `field_notes.csv` row. Coordinate flags become data_flags."""
    row = {column: '' for column in FIELD_NOTES_COLUMNS}
    flags: list[str] = []

    for field in visit_schema.VISIT_FIELDS:
        raw = values.get(field.column)
        column = field.column

        if column in ('lat', 'lon'):
            value, flag = visit_schema.read_coordinate(raw, column)
            row[column] = '' if value is None else f'{value:.5f}'
            if flag:
                flags.append(flag)
        elif field.options:
            row[column] = _read_choice(raw, field, where, problems)
        elif field.bounds:
            row[column] = _read_number(raw, field, where, problems)
        elif column == 'visit_date':
            row[column] = _read_date(raw, column, where, problems)
        elif column == 'visit_time':
            row[column] = _read_time(raw, column, where, problems)
        elif column in ('camera_datetime_observed', 'camera_datetime_after'):
            row[column] = _read_datetime(raw, column, where, problems)
        else:
            row[column] = _text(raw)

    _check_obligations(row, where, problems)
    row['source_sheet'] = source_sheet
    row['data_flags'] = '; '.join(flags)
    return row


# =============================================================================
# The two public verbs
# =============================================================================

def read(workbook: Path) -> list[dict]:
    """Every filled visit in `workbook`, validated, in sheet order.

    Raises `VisitFormError` listing every problem. An empty sheet is not a problem:
    a salida that recorded nothing returns an empty list.
    """
    # Imported here, not at module scope: `timestamps.py` pulls in this package for
    # `anchors` on every ingest and has no reason to need an Excel reader.
    from openpyxl import load_workbook

    workbook = Path(workbook)
    wb = load_workbook(workbook, data_only=True)
    if SHEET not in wb.sheetnames:
        raise VisitFormError(workbook, [
            f'la planilla no tiene hoja {SHEET!r} (tiene: {", ".join(wb.sheetnames)})'])
    ws = wb[SHEET]

    problems: list[str] = []
    columns: dict[int, str] = {}
    for index, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        if _blank(cell.value):
            continue
        try:
            columns[index] = visit_schema.by_label(str(cell.value)).column
        except KeyError as exc:
            problems.append(str(exc.args[0]))

    missing = [f.column for f in visit_schema.VISIT_FIELDS
               if f.column not in columns.values()]
    if missing:
        problems.append(f'faltan columnas en la hoja: {", ".join(missing)}')
    if problems:
        raise VisitFormError(workbook, problems)

    source_sheet = f'{SHEET} ({workbook.name})'
    rows: list[dict] = []
    for number, cells in enumerate(ws.iter_rows(min_row=2), start=2):
        values = {columns[i]: c.value for i, c in enumerate(cells, start=1)
                  if i in columns}
        if all(_blank(v) for v in values.values()):
            continue
        rows.append(_read_row(values, f'fila {number}', source_sheet, problems))

    _refuse_duplicates(rows, problems)
    if problems:
        raise VisitFormError(workbook, problems)
    return rows


def ingest(workbook: Path, csv_path: Path | None = None) -> int:
    """Append `workbook`'s visits to `field_notes.csv`. Returns rows appended.

    Idempotent: re-running the same workbook refuses rather than duplicating, so
    "did that run land?" is answered by running it again.
    """
    workbook = Path(workbook)
    csv_path = Path(csv_path) if csv_path else _default_csv(workbook)
    rows = read(workbook)
    if not rows:
        return 0

    existing = _existing_rows(csv_path)
    known = {_key(r) for r in existing}
    clashes = [f'{_key(r)} ya está en {csv_path.name}' for r in rows if _key(r) in known]
    if clashes:
        raise VisitFormError(workbook, clashes)

    is_new = not csv_path.exists()
    with csv_path.open('a', encoding='utf-8', newline='') as fh:
        writer = csv.DictWriter(fh, fieldnames=list(FIELD_NOTES_COLUMNS))
        if is_new:
            writer.writeheader()
        writer.writerows(rows)
    return len(rows)


# =============================================================================
# Helpers
# =============================================================================

def _default_csv(workbook: Path) -> Path:
    """The record beside the workbook. Both live in `data/campaigns/`."""
    return workbook.parent / FIELD_NOTES_FILENAME


def _key(row: dict) -> tuple:
    return tuple(row.get(column, '') for column in VISIT_KEY)


def _refuse_duplicates(rows: list[dict], problems: list[str]) -> None:
    seen: dict[tuple, int] = {}
    for number, row in enumerate(rows, start=2):
        key = _key(row)
        if key in seen:
            problems.append(f'visita repetida en la planilla: {key}')
        seen[key] = number


def _existing_rows(csv_path: Path) -> list[dict]:
    if not csv_path.exists():
        return []
    with csv_path.open(encoding='utf-8', newline='') as fh:
        reader = csv.DictReader(fh)
        fieldnames = reader.fieldnames or []
        if RECORDED_CLOSE_COLUMN in fieldnames:
            raise VisitFormError(csv_path, [
                f'{csv_path.name} todavía tiene la columna retirada '
                f'`{RECORDED_CLOSE_COLUMN}`: es una copia previa al cambio de forma. '
                'Ejecutar setup/reshape_field_notes.py antes de cargar visitas.'])
        return list(reader)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument('workbook', type=Path, help='Registro de visitas CT.xlsx (filled)')
    ap.add_argument('--csv', type=Path, default=None,
                    help=f'target {FIELD_NOTES_FILENAME} (default: beside the workbook)')
    ap.add_argument('--check', action='store_true',
                    help='validate and report; write nothing')
    args = ap.parse_args(argv)

    try:
        if args.check:
            rows = read(args.workbook)
            print(f'{len(rows)} visita(s) válida(s). No se escribió nada.')
            return 0
        appended = ingest(args.workbook, args.csv)
    except VisitFormError as exc:
        print(exc)
        return 1

    print(f'{appended} visita(s) agregada(s) a '
          f'{args.csv or _default_csv(Path(args.workbook))}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
