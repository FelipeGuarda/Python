"""The structure of the canonical contact list.

This module owns what the master workbook *is*: which columns mean what, that
addresses may be packed several to a cell, how the `N` column continues, that
the autofilter has to grow with the data, and that the file is a colleague's
working document whose formatting — his highlights above all — must survive
every write. Since rows may now be corrected and consolidated as well as
appended, it also owns what a write is allowed to destroy: nothing. A merge
moves what it cannot keep in place into `Notas`, and a deletion happens only
after every read the plan depends on.

The originating file is never opened for writing. Every operation produces a
new file and returns its path.
"""

from __future__ import annotations

import difflib
import re
import unicodedata
from copy import copy
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .rosters import extract_emails

# Canonical column order. Position is meaningful: the first four are what the
# list has always had, so restructuring only ever appends to the right and
# never disturbs an existing cell.
COLUMNS = [
    "N",
    "Nombre",
    "Organización",
    "Email principal",
    "Email alternativo",
    "Consentimiento",
    "Origen",
    "Fecha",
    "Notas",
]

# What gets shared. Regenerated from the canonical file, never hand-edited.
SHARE_COLUMNS = ["N", "Nombre", "Organización", "Email principal", "Email alternativo"]

# The header the list used before restructuring, when one cell held every
# address a person had.
LEGACY_EMAIL_HEADER = "Email(s)"

WIDTHS = {
    "N": 6, "Nombre": 30.4, "Organización": 35, "Email principal": 38,
    "Email alternativo": 38, "Consentimiento": 15, "Origen": 34,
    "Fecha": 13, "Notas": 40,
}

HEADER_ROW = 1
FIRST_DATA_ROW = 2


@dataclass
class Restructured:
    """What changed when the sheet was brought up to the canonical columns."""

    split_addresses: list[tuple[int, str, list[str]]] = field(default_factory=list)
    numbers_frozen: int = 0


@dataclass(frozen=True)
class Person:
    row: int
    nombre: str
    organizacion: str
    addresses: list[str]
    # `N` as displayed, which is not the row: the owner sorts by name. Carried
    # here so callers can reason about join order without knowing that reading
    # it means going through a second, data-only read of the workbook.
    numero: int | None = None


@dataclass(frozen=True)
class Amendment:
    """A correction to a row that is already on the list.

    `None` leaves a field alone; `""` clears it. Notas are appended to
    whatever the row already says rather than replacing it — the owner's note
    and a cargo lifted out of `Organización` both belong there.
    """

    row: int
    organizacion: str | None = None
    notas: str | None = None


@dataclass(frozen=True)
class Merge:
    """Two rows that are one person.

    `keep` survives and gains every address both rows held; `drop` is removed.
    Which of the two to keep is a judgement about which organisation is
    current, so it is decided by the reviewer, not here.
    """

    keep: int
    drop: int
    organizacion: str | None = None
    notas: str | None = None


@dataclass
class Curation:
    """Everything to change about rows already on the list, in one object.

    Handed over whole because applying it in the wrong order corrupts the
    sheet: a deletion shifts every row beneath it, so row numbers taken from
    a review sheet stop meaning what they meant. `MasterList.curate` owns
    that ordering and callers never see it.
    """

    amendments: list[Amendment] = field(default_factory=list)
    merges: list[Merge] = field(default_factory=list)


@dataclass
class CurationReport:
    """What curating actually changed, for the operator to read back."""

    organizaciones: list[tuple[str, str, str]] = field(default_factory=list)
    notas: int = 0
    addresses_added: list[tuple[str, str]] = field(default_factory=list)
    rows_removed: list[tuple[int, str]] = field(default_factory=list)
    carried_over: list[tuple[str, str, str]] = field(default_factory=list)


@dataclass
class NewContact:
    """A contact staged for appending."""

    nombre: str
    organizacion: str = ""
    email_principal: str = ""
    email_alternativo: str = ""
    consentimiento: str = ""
    notas: str = ""

    @property
    def addresses(self) -> list[str]:
        return [a for a in (self.email_principal, self.email_alternativo) if a]


def _fold(text: str) -> str:
    """Name reduced for comparison: accent-free, lowercase, letters only."""
    folded = unicodedata.normalize("NFKD", str(text or "").lower())
    folded = "".join(c for c in folded if not unicodedata.combining(c))
    return re.sub(r"[^a-z ]", " ", folded).strip()


def _copy_style(source, target) -> None:
    """Carry a cell's look forward — but never its fill.

    Fills are how the list's owner marks rows for his own attention.
    Propagating one onto an appended row would invent a highlight he did not
    make; leaving it off keeps highlighting his to give.
    """
    target.font = copy(source.font)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format


class MasterList:
    """The canonical contact workbook."""

    def __init__(self, path: Path, workbook, sheet: Worksheet, cached: Worksheet | None):
        self.path = path
        self.workbook = workbook
        self.sheet = sheet
        self._cached = cached

    @classmethod
    def open(cls, path: str | Path, sheet: str | None = None) -> "MasterList":
        path = Path(path)
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook[sheet] if sheet else workbook.worksheets[0]
        # A second read resolves the `N` column's formulas to their values;
        # openpyxl cannot give both formulas and results from one load.
        cached_book = openpyxl.load_workbook(path, data_only=True)
        cached = cached_book[worksheet.title]
        return cls(path, workbook, worksheet, cached)

    @property
    def headers(self) -> list[str]:
        return [
            str(self.sheet.cell(HEADER_ROW, c).value or "").strip()
            for c in range(1, self.sheet.max_column + 1)
        ]

    @property
    def is_restructured(self) -> bool:
        return self.headers[:len(COLUMNS)] == COLUMNS

    def column(self, name: str) -> int:
        """1-based index of a canonical column."""
        return COLUMNS.index(name) + 1

    @property
    def last_data_row(self) -> int:
        return self.sheet.max_row

    def _email_column(self) -> int:
        headers = self.headers
        for index, header in enumerate(headers, start=1):
            if header in ("Email principal", LEGACY_EMAIL_HEADER):
                return index
        return 4

    @property
    def people(self) -> list[Person]:
        email_col = self._email_column()
        alt_col = self.column("Email alternativo") if self.is_restructured else None
        found = []
        for row in range(FIRST_DATA_ROW, self.last_data_row + 1):
            nombre = str(self.sheet.cell(row, 2).value or "").strip()
            addresses = extract_emails(self.sheet.cell(row, email_col).value)
            if alt_col:
                addresses += extract_emails(self.sheet.cell(row, alt_col).value)
            # An organisation that registered without naming a person still
            # occupies a row and still owns its address — skipping it here
            # would let the same address be re-added as new next time.
            if not nombre and not addresses:
                continue
            numero = self._cached_value(row, 1)
            found.append(Person(
                row=row,
                nombre=nombre,
                organizacion=str(self.sheet.cell(row, 3).value or "").strip(),
                addresses=addresses,
                numero=int(numero) if isinstance(numero, (int, float)) else None,
            ))
        return found

    @property
    def addresses(self) -> set[str]:
        return {a for person in self.people for a in person.addresses}

    def similar_names(self, name: str, cutoff: float = 0.86) -> list[Person]:
        """People already on the list whose name plausibly matches `name`.

        This is what catches the same person arriving under a second address —
        which matching on the address alone cannot see.

        Two tests, because one is not enough. Edit distance catches spelling
        drift ("Nelida"/"Nélida"). Token containment catches the surname that
        appears on one form and not the next: Spanish names carry one or two
        surnames inconsistently, and "Natasha Pons" against "Natasha Pons
        Majmut" scores only 0.77 — under any cutoff loose enough to be safe.
        """
        target = _fold(name)
        tokens = set(target.split())
        if len(tokens) < 2:
            return []  # a single token is too weak to match on

        matches: dict[int, Person] = {}
        for person in self.people:
            folded = _fold(person.nombre)
            other = set(folded.split())
            if len(other) < 2:
                continue
            subset = tokens <= other or other <= tokens
            if (subset and len(tokens & other) >= 2) or (
                difflib.SequenceMatcher(None, target, folded).ratio() >= cutoff
            ):
                matches[person.row] = person
        return list(matches.values())[:3]

    def rows_without_address(self) -> list[Person]:
        """Rows whose address cell holds something other than an address.

        These are left exactly as they are by every operation here — the cell
        content and any highlight on it belong to the list's owner.
        """
        return [p for p in self.people if not p.addresses]

    def restructure(self) -> Restructured:
        """Bring the sheet up to the canonical column set, in memory.

        Existing cells are untouched apart from splitting the packed address
        cells: the first address stays where it is, the second moves to
        `Email alternativo`. New columns are added to the right, so no cell
        ever shifts position and no formatting moves with it.

        Returns the rows whose addresses were split, for reporting. Call
        `save_as` to persist.
        """
        sheet = self.sheet
        template = sheet.cell(HEADER_ROW, 4)  # an existing header, for its look

        report = Restructured()
        if not self.is_restructured:
            report.numbers_frozen = self._freeze_numbering()
            for index, name in enumerate(COLUMNS, start=1):
                cell = sheet.cell(HEADER_ROW, index)
                if index == 4 or cell.value in (None, ""):
                    cell.value = name
                if index > 4:
                    _copy_style(template, cell)
                    cell.fill = copy(template.fill)  # headers do carry their fill

            email_col, alt_col = 4, self.column("Email alternativo")
            for row in range(FIRST_DATA_ROW, self.last_data_row + 1):
                found = extract_emails(sheet.cell(row, email_col).value)
                if len(found) < 2:
                    continue  # one address, or a note the owner left — leave it
                source = sheet.cell(row, email_col)
                sheet.cell(row, alt_col).value = "; ".join(found[1:])
                _copy_style(source, sheet.cell(row, alt_col))
                source.value = found[0]
                report.split_addresses.append((row, found[0], found[1:]))

        self._apply_widths()
        self._refresh_autofilter()
        return report

    def _freeze_numbering(self) -> int:
        """Replace `N`'s formulas with the numbers they already evaluate to.

        openpyxl discards cached formula results when it saves, so a `=A2+1`
        left in place reads back as empty for anything downstream — the shared
        copy included. The displayed numbers do not change, and the column
        ends up consistent with rows 120 onward, which the list's owner had
        already pasted as literals.
        """
        frozen = 0
        for row in range(FIRST_DATA_ROW, self.last_data_row + 1):
            cell = self.sheet.cell(row, 1)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            cached = self._cached_value(row, 1)
            if isinstance(cached, (int, float)):
                cell.value = int(cached)
                frozen += 1
        return frozen

    def curate(self, plan: Curation) -> CurationReport:
        """Correct and consolidate rows that are already on the list.

        This is the only operation here that alters and removes the owner's
        rows, so it does the least it can: it writes only the fields the plan
        names, never touches a fill, and before removing a row it carries
        every value that row held alone onto the survivor. Nothing the owner
        typed is lost by a merge — at worst it moves to `Notas`.

        Requires the canonical columns: uniting two people's addresses needs
        somewhere to put the second one. Call `restructure` first.
        """
        if not self.is_restructured:
            raise ValueError(
                "curate needs the canonical columns — call restructure() first, "
                "or a merged address has nowhere to go."
            )

        report = CurationReport()
        for amendment in plan.amendments:
            self._amend(amendment, report)

        # Every read and write finishes before any row is deleted. Deleting
        # shifts everything beneath it, which would silently repoint the row
        # numbers the rest of the plan still refers to.
        for merge in plan.merges:
            self._fuse(merge, report)
        for row, _ in sorted(report.rows_removed, reverse=True):
            self.sheet.delete_rows(row)
            if self._cached:
                self._cached.delete_rows(row)

        self._apply_widths()
        self._refresh_autofilter()
        return report

    def _amend(self, amendment: Amendment, report: CurationReport) -> None:
        row = amendment.row
        nombre = str(self.sheet.cell(row, 2).value or "")
        if amendment.organizacion is not None:
            cell = self.sheet.cell(row, self.column("Organización"))
            before, after = str(cell.value or "").strip(), amendment.organizacion.strip()
            if before != after:
                cell.value = after or None
                report.organizaciones.append((nombre, before, after))
        if amendment.notas and self._append_note(row, amendment.notas):
            report.notas += 1

    def _append_note(self, row: int, text: str) -> bool:
        """Add to a row's notes without displacing what is already there."""
        cell = self.sheet.cell(row, self.column("Notas"))
        existing = str(cell.value or "").strip()
        if existing.lower() == "nan":
            existing = ""  # a pandas NaN that reached the sheet as text
        if text in existing:
            return False
        cell.value = f"{existing}; {text}" if existing else text
        return True

    def _fuse(self, merge: Merge, report: CurationReport) -> None:
        """Union one row into another and mark the loser for removal."""
        keep, drop = merge.keep, merge.drop
        principal, alternativo = self.column("Email principal"), self.column("Email alternativo")
        nombre = str(self.sheet.cell(keep, 2).value or "")

        held = extract_emails(self.sheet.cell(keep, principal).value)
        held += extract_emails(self.sheet.cell(keep, alternativo).value)
        incoming = extract_emails(self.sheet.cell(drop, principal).value)
        incoming += extract_emails(self.sheet.cell(drop, alternativo).value)
        arriving = [a for a in dict.fromkeys(incoming) if a not in set(held)]

        # The survivor's own address cell is left exactly as written — the
        # owner's capitalisation of it is not ours to normalise.
        if not held and arriving:
            self.sheet.cell(keep, principal).value = arriving.pop(0)
        if arriving:
            existing = str(self.sheet.cell(keep, alternativo).value or "").strip()
            parts = ([existing] if existing else []) + arriving
            self.sheet.cell(keep, alternativo).value = "; ".join(parts)
        for address in arriving:
            report.addresses_added.append((nombre, address))

        for name in ("Consentimiento", "Origen", "Fecha"):
            column = self.column(name)
            surviving, losing = self.sheet.cell(keep, column), self.sheet.cell(drop, column)
            if surviving.value in (None, "") and losing.value not in (None, ""):
                surviving.value = losing.value
                surviving.number_format = losing.number_format
                report.carried_over.append((nombre, name, str(losing.value)))

        # The other row's organisation is the person's history, not a mistake:
        # Stowhas moved ministries. It goes to Notas rather than nowhere.
        theirs = str(self.sheet.cell(drop, self.column("Organización")).value or "").strip()
        ours = str(self.sheet.cell(keep, self.column("Organización")).value or "").strip()
        if theirs and theirs.casefold() != ours.casefold():
            self._append_note(keep, f"antes: {theirs}")
        their_note = str(self.sheet.cell(drop, self.column("Notas")).value or "").strip()
        if their_note and their_note.lower() != "nan":
            self._append_note(keep, their_note)

        if merge.organizacion is not None or merge.notas is not None:
            self._amend(Amendment(keep, merge.organizacion, merge.notas), report)

        report.rows_removed.append((drop, str(self.sheet.cell(drop, 2).value or "")))

    def append(
        self,
        contacts: list[NewContact],
        origen: str,
        fecha: date,
    ) -> None:
        """Add contacts to the bottom of the list, in the order given.

        Nothing is sorted and no existing row is touched. `Origen` and `Fecha`
        are stamped here rather than read from the source, and are written
        only on the rows being added — an existing row's provenance records
        when that person joined and is never rewritten.

        Call `save_as` to persist.
        """
        sheet = self.sheet
        style_row = self.last_data_row
        next_number = self._next_number()

        for offset, contact in enumerate(contacts):
            row = self.last_data_row + 1
            values = {
                "N": next_number + offset,
                "Nombre": contact.nombre,
                "Organización": contact.organizacion,
                "Email principal": contact.email_principal,
                "Email alternativo": contact.email_alternativo,
                "Consentimiento": contact.consentimiento,
                "Origen": origen,
                "Fecha": fecha,
                "Notas": contact.notas,
            }
            for name, value in values.items():
                index = self.column(name)
                cell = sheet.cell(row, index)
                cell.value = value if value not in ("", None) else None
                _copy_style(sheet.cell(style_row, index), cell)
                if name == "Fecha":
                    cell.number_format = "yyyy-mm-dd"

        self._apply_widths()
        self._refresh_autofilter()

    def export_share(self, to: str | Path) -> Path:
        """Write the trimmed copy meant for circulation.

        Generated, never edited by hand: an edit made here and not in the
        canonical file is an edit that will be silently overwritten.
        """
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = self.sheet.title

        header_style = self.sheet.cell(HEADER_ROW, 1)
        for index, name in enumerate(SHARE_COLUMNS, start=1):
            cell = sheet.cell(HEADER_ROW, index, name)
            _copy_style(header_style, cell)
            cell.fill = copy(header_style.fill)
            sheet.column_dimensions[get_column_letter(index)].width = WIDTHS.get(name, 20)

        source_columns = [self.column(n) for n in SHARE_COLUMNS]
        out_row = FIRST_DATA_ROW
        for row in range(FIRST_DATA_ROW, self.last_data_row + 1):
            if not any(self.sheet.cell(row, c).value for c in source_columns):
                continue  # only genuinely empty rows are dropped
            for index, source_column in enumerate(source_columns, start=1):
                value = self.sheet.cell(row, source_column).value
                if isinstance(value, str) and value.startswith("="):
                    value = self._cached_value(row, source_column)
                sheet.cell(out_row, index, value)
            out_row += 1

        sheet.auto_filter.ref = f"A1:{get_column_letter(len(SHARE_COLUMNS))}{out_row - 1}"
        to = Path(to)
        to.parent.mkdir(parents=True, exist_ok=True)
        book.save(to)
        return to

    def _cached_value(self, row: int, column: int):
        return self._cached.cell(row, column).value if self._cached else None

    def _next_number(self) -> int:
        """The next value for `N`, one past the highest already in use.

        The column mixes formulas (early rows) with pasted literals (later
        ones), so the cached result is read rather than the cell content.

        It is the maximum rather than the last row's value because `N` records
        the order people joined, not where they sit: the list's owner sorts the
        sheet by name, which scatters the numbers, and adds rows without
        numbering them. Reading the bottom row gave 130 against a list whose
        highest number was 141, which would have re-used a dozen numbers.
        """
        used = [
            self._cached_value(row, 1)
            for row in range(FIRST_DATA_ROW, self.last_data_row + 1)
        ]
        numbers = [int(v) for v in used if isinstance(v, (int, float))]
        return max(numbers) + 1 if numbers else 1

    def fill_missing_numbers(self) -> list[tuple[int, int]]:
        """Number any row the owner added without one.

        Returns the (row, number) pairs assigned, top to bottom. Idempotent:
        a second call finds nothing to do.
        """
        assigned = []
        for row in range(FIRST_DATA_ROW, self.last_data_row + 1):
            if self._cached_value(row, 1) is not None:
                continue
            if not self.sheet.cell(row, 2).value and not self.sheet.cell(row, 3).value:
                continue  # a genuinely empty row is not a person
            number = self._next_number()
            self.sheet.cell(row, 1).value = number
            if self._cached:
                self._cached.cell(row, 1).value = number
            _copy_style(self.sheet.cell(row - 1, 1), self.sheet.cell(row, 1))
            assigned.append((row, number))
        return assigned

    def _apply_widths(self) -> None:
        for index, name in enumerate(COLUMNS, start=1):
            letter = get_column_letter(index)
            if name in WIDTHS and not self.sheet.column_dimensions[letter].width:
                self.sheet.column_dimensions[letter].width = WIDTHS[name]

    def _refresh_autofilter(self) -> None:
        """Grow the filter to cover every column and row, or new rows sit
        outside it and quietly vanish whenever someone filters."""
        last_column = get_column_letter(max(self.sheet.max_column, len(COLUMNS)))
        self.sheet.auto_filter.ref = f"A1:{last_column}{self.last_data_row}"

    def save_as(self, to: str | Path) -> Path:
        to = Path(to)
        if to.resolve() == self.path.resolve():
            raise ValueError(
                f"Refusing to overwrite the canonical file {self.path.name}. "
                "Write to a new path and replace it yourself once you are satisfied."
            )
        to.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(to)
        return to
