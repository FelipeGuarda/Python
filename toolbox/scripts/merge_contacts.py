"""Merge event contact lists into the canonical master, in two passes.

Pass one proposes; you correct it in Excel; pass two appends.

    python scripts/merge_contacts.py --review  --master M.xlsx --source A.xlsx --source B.xlsx
    # edit revision.xlsx: fix any split, set Añadir=NO to reject
    python scripts/merge_contacts.py --apply   --master M.xlsx \
        --origen "3° Encuentro Hablemos de Conservación" --fecha 2026-07-28

The split between passes exists because the sources pack a person and their
organisation into one free-text cell, and some of those cannot be taken apart
by any rule. Rather than guess, the doubtful ones are flagged for a human.

The master is never written to. Both passes produce new files.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from lib.master_list import MasterList, NewContact  # noqa: E402
from lib.namesplit import split_contact  # noqa: E402
from lib.rosters import KEY, extract_emails, load_roster  # noqa: E402

REVIEW_COLUMNS = [
    "Añadir", "Nombre", "Organización", "Email principal", "Email alternativo",
    "Consentimiento", "Confianza", "Motivo", "Original", "Fuente",
    "Posible duplicado", "Notas",
]

# Review sheet cues. Amber marks a split the rules could not settle; red marks
# someone who looks like they are already on the list under another address.
FILL_BAJA = PatternFill("solid", fgColor="FFF2CC")
FILL_DUPLICADO = PatternFill("solid", fgColor="FCE4D6")
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")

CONFIDENCE_ORDER = {"baja": 0, "media": 1, "alta": 2}


def _find_column(frame, *needles: str) -> str | None:
    """First column whose header contains any of `needles`, case-folded."""
    for column in frame.columns:
        folded = str(column).lower()
        if any(needle in folded for needle in needles):
            return column
    return None


def _collect(source: Path, known: set[str], name_column: str | None) -> list[dict]:
    """Read one source file into candidate records, skipping known addresses."""
    roster = load_roster(source)
    frame = roster.unique
    packed_column = name_column or _find_column(frame, "nombre", "name")
    if packed_column is None:
        raise SystemExit(
            f"{source.name}: no column looks like a name. "
            f"Columns are: {', '.join(str(c) for c in frame.columns if c != KEY)}. "
            f"Pass --name-col to say which one."
        )
    notes_column = _find_column(frame, "nota", "observ")

    candidates = []
    for _, row in frame.iterrows():
        addresses = extract_emails(row[roster.key_column])
        if not addresses or any(a in known for a in addresses):
            continue
        split = split_contact(row[packed_column])
        candidates.append({
            "addresses": addresses,
            "split": split,
            "notas": str(row[notes_column]).strip() if notes_column and row[notes_column] else "",
            "fuente": f"{source.stem[:28]}",
        })
    return candidates


def _merge_duplicates(candidates: list[dict]) -> list[dict]:
    """Fold records sharing an address into one, keeping the clearest split."""
    merged: dict[str, dict] = {}
    for candidate in candidates:
        key = candidate["addresses"][0]
        existing = merged.get(key)
        if existing is None:
            merged[key] = candidate
            continue
        for address in candidate["addresses"]:
            if address not in existing["addresses"]:
                existing["addresses"].append(address)
        better = CONFIDENCE_ORDER[candidate["split"].confianza]
        if better > CONFIDENCE_ORDER[existing["split"].confianza]:
            existing["split"] = candidate["split"]
        if candidate["notas"] and not existing["notas"]:
            existing["notas"] = candidate["notas"]
        if candidate["fuente"] not in existing["fuente"]:
            existing["fuente"] += f" + {candidate['fuente']}"
    return list(merged.values())


def review(args: argparse.Namespace) -> int:
    master = MasterList.open(args.master)
    known = master.addresses
    print(f"Master: {len(master.people)} personas, {len(known)} direcciones "
          f"({'reestructurado' if master.is_restructured else 'formato antiguo'})")

    candidates: list[dict] = []
    for source in args.source:
        found = _collect(source, known, args.name_col)
        print(f"  {source.name}: {len(found)} direcciones nuevas")
        candidates.extend(found)

    records = _merge_duplicates(candidates)
    print(f"\n{len(records)} contactos nuevos tras unir repetidos entre archivos")

    for record in records:
        split = record["split"]
        record["duplicados"] = master.similar_names(split.nombre) if split.nombre else []

    # Probable duplicates first, then least-confident splits: the sheet opens
    # on whatever most needs a decision.
    records.sort(key=lambda r: (
        not r["duplicados"],
        CONFIDENCE_ORDER[r["split"].confianza],
        r["split"].nombre.lower(),
    ))

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Revisión"
    for index, name in enumerate(REVIEW_COLUMNS, start=1):
        cell = sheet.cell(1, index, name)
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    flagged = doubtful = 0
    for offset, record in enumerate(records):
        row = offset + 2
        split = record["split"]
        duplicates = record["duplicados"]
        addresses = record["addresses"]
        matches = "; ".join(
            f"fila {p.row}: {p.nombre} <{p.addresses[0] if p.addresses else 'sin email'}>"
            for p in duplicates
        )
        values = [
            # A suspected duplicate defaults to NO: adding one is a mistake
            # that is hard to see later, skipping one is visible immediately.
            "NO" if duplicates else "SI",
            split.nombre,
            split.organizacion,
            addresses[0],
            "; ".join(addresses[1:]),
            "",  # these forms never asked for consent
            split.confianza,
            split.motivo,
            split.original,
            record["fuente"],
            matches,
            record["notas"],
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row, index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=index in (8, 9, 11))
            if duplicates:
                cell.fill = FILL_DUPLICADO
            elif split.confianza == "baja":
                cell.fill = FILL_BAJA
        flagged += bool(duplicates)
        doubtful += split.confianza == "baja" and not duplicates

    for index, name in enumerate(REVIEW_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = {
            "Añadir": 8, "Motivo": 34, "Original": 38, "Posible duplicado": 46,
        }.get(name, 24)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_COLUMNS))}{len(records) + 1}"

    output = Path(args.review_file)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)

    print(f"  {flagged} marcados como posible duplicado (Añadir=NO por defecto)")
    print(f"  {doubtful} con separación dudosa (confianza baja)")
    print(f"  {len(records) - flagged - doubtful} sin observaciones")
    print(f"\nRevisa y corrige: {output}")
    print("Luego: --apply con --origen y --fecha")
    return 0


def _read_review(path: Path) -> list[NewContact]:
    sheet = openpyxl.load_workbook(path).active
    headers = [str(c.value or "").strip() for c in sheet[1]]
    index = {name: position for position, name in enumerate(headers)}
    missing = [c for c in ("Añadir", "Nombre", "Email principal") if c not in index]
    if missing:
        raise SystemExit(f"{path.name}: faltan columnas {missing}. ¿Es el archivo de revisión?")

    def value(row, name):
        position = index.get(name)
        if position is None or not row[position].value:
            return ""
        text = str(row[position].value).strip()
        # An empty Notas cell round-trips through pandas as the string "nan".
        # Left alone it reaches the master as a note reading "nan".
        return "" if text.lower() == "nan" else text

    contacts, rejected = [], 0
    for row in sheet.iter_rows(min_row=2):
        if not any(c.value for c in row):
            continue
        if value(row, "Añadir").upper() not in ("SI", "SÍ", "S", "YES", "Y"):
            rejected += 1
            continue
        contacts.append(NewContact(
            nombre=value(row, "Nombre"),
            organizacion=value(row, "Organización"),
            email_principal=value(row, "Email principal"),
            email_alternativo=value(row, "Email alternativo"),
            consentimiento=value(row, "Consentimiento"),
            notas=value(row, "Notas"),
        ))
    print(f"Revisión: {len(contacts)} para añadir, {rejected} descartados")
    return contacts


def apply(args: argparse.Namespace) -> int:
    review_path = Path(args.review_file)
    if not review_path.exists():
        raise SystemExit(f"No existe {review_path}. Ejecuta primero --review.")
    contacts = _read_review(review_path)
    if not contacts:
        print("Nada que añadir.")
        return 0

    master = MasterList.open(args.master)
    if not master.is_restructured:
        report = master.restructure()
        print(f"Reestructurado a {len(master.headers)} columnas; "
              f"{len(report.split_addresses)} celdas con dos direcciones separadas; "
              f"{report.numbers_frozen} fórmulas de N convertidas a su valor")
    left_alone = master.rows_without_address()
    if left_alone:
        print(f"{len(left_alone)} fila(s) sin dirección quedan intactas: "
              + ", ".join(f"fila {p.row} ({p.nombre})" for p in left_alone))

    before = master.last_data_row
    master.append(contacts, origen=args.origen, fecha=args.fecha)
    output = master.save_as(args.output)
    print(f"\nAñadidas {len(contacts)} filas ({before + 1}–{master.last_data_row}) -> {output}")

    share = MasterList.open(output).export_share(args.share_file)
    print(f"Copia para compartir -> {share}")
    print("\nEl archivo original no fue modificado. Revísalo y reemplázalo tú.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Añade contactos de formularios de eventos a la lista maestra.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--master", type=Path, required=True, help="lista maestra canónica")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--review", action="store_true", help="proponer (no escribe la maestra)")
    mode.add_argument("--apply", action="store_true", help="añadir lo aprobado en la revisión")

    parser.add_argument("--source", type=Path, action="append", default=[],
                        help="archivo de origen (repetible); sólo con --review")
    parser.add_argument("--name-col", help="forzar la columna con nombre+organización")
    parser.add_argument("--review-file", type=Path, default=Path("revision.xlsx"),
                        help="hoja de revisión (por defecto: revision.xlsx)")
    parser.add_argument("--origen", help="nombre del encuentro; obligatorio con --apply")
    parser.add_argument("--fecha", type=lambda s: datetime.strptime(s, "%Y-%m-%d").date(),
                        help="fecha del encuentro AAAA-MM-DD; obligatoria con --apply")
    parser.add_argument("--output", type=Path, help="maestra actualizada (por defecto: *_actualizado.xlsx)")
    parser.add_argument("--share-file", type=Path, help="copia para compartir (por defecto: *_COMPARTIR.xlsx)")
    args = parser.parse_args()

    if not args.master.exists():
        parser.error(f"no existe: {args.master}")

    if args.review:
        if not args.source:
            parser.error("--review necesita al menos un --source")
        for source in args.source:
            if not source.exists():
                parser.error(f"no existe: {source}")
        return review(args)

    if not args.origen or not args.fecha:
        parser.error("--apply necesita --origen y --fecha (no se deducen del archivo)")
    stem = args.master.stem
    args.output = args.output or args.master.with_name(f"{stem}_actualizado.xlsx")
    args.share_file = args.share_file or args.master.with_name(f"{stem}_COMPARTIR.xlsx")
    return apply(args)


if __name__ == "__main__":
    raise SystemExit(main())
