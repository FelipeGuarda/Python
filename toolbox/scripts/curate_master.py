"""Correct the rows already on the master list, in two passes.

Pass one proposes; you correct it in Excel; pass two applies.

    python scripts/curate_master.py --review --master M.xlsx
    # edit curacion.xlsx: fix an organisation, set Aplicar=NO to skip
    python scripts/curate_master.py --apply  --master M.xlsx

Two kinds of correction, because two kinds of mess accumulate in a list people
fill in by hand:

- **cargo** — a job title typed into `Organización` ("Directora Educación
  MIM"). The title moves to `Notas`; the organisation stays behind.
- **fusionar** — one person occupying two rows, usually because they
  registered once with a personal address and once with an institutional one.
  The rows become one, holding both addresses.

Where the merge script asks *is this person new?*, this one asks *is this row
right?* — so it proposes and lets a human settle it, for the same reason: the
rules can see that a cell is wrong far more reliably than they can see what it
should say instead.

The master is never written to. Both passes produce new files.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from lib.master_list import Amendment, Curation, MasterList, Merge  # noqa: E402
from lib.namesplit import split_cargo  # noqa: E402

REVIEW_COLUMNS = [
    "Aplicar", "Acción", "Fila", "Nombre", "Organización actual",
    "Organización propuesta", "Cargo a Notas", "Confianza", "Motivo",
    "Fila a eliminar", "Se fusiona con", "Direcciones",
]

# Same cues as revision.xlsx, so one glance means the same thing in both:
# amber for a proposal the rules could not settle, red for a row disappearing.
FILL_DUDOSO = PatternFill("solid", fgColor="FFF2CC")
FILL_FUSION = PatternFill("solid", fgColor="FCE4D6")
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")

CONFIDENCE_ORDER = {"baja": 0, "media": 1, "alta": 2}
WIDTHS = {
    "Aplicar": 8, "Acción": 10, "Fila": 6, "Nombre": 26,
    "Organización actual": 34, "Organización propuesta": 30,
    "Cargo a Notas": 30, "Confianza": 10, "Motivo": 30,
    "Fila a eliminar": 14, "Se fusiona con": 40, "Direcciones": 46,
}
YES = ("SI", "SÍ", "S", "YES", "Y")


def _proposed_merges(master: MasterList) -> list[dict]:
    """Rows that are the same person, and which of the two should survive.

    The survivor is the row whose `Organización` names an organisation rather
    than a job title; where both do, the higher `N` wins, because `N` records
    the order people joined and the later row carries the current employer.
    Both halves are shown in the sheet so a reviewer can swap them.
    """
    people = {person.row: person for person in master.people}
    pairs: dict[tuple[int, int], tuple] = {}
    for person in people.values():
        if not person.nombre:
            continue
        for other in master.similar_names(person.nombre):
            if other.row == person.row or not other.nombre:
                continue
            pairs.setdefault(tuple(sorted((person.row, other.row))), (person, other))

    def informative(candidate) -> tuple:
        names_an_org = bool(candidate.organizacion) and not split_cargo(candidate.organizacion).has_cargo
        return (names_an_org, candidate.numero or 0)

    proposals = []
    for (left_row, right_row), _ in sorted(pairs.items()):
        left, right = people[left_row], people[right_row]
        keep, drop = (left, right) if informative(left) >= informative(right) else (right, left)
        shared = set(keep.addresses) & set(drop.addresses)
        proposals.append({
            "keep": keep,
            "drop": drop,
            "motivo": "misma dirección en ambas filas" if shared else "mismo nombre, direcciones distintas",
            "confianza": "alta" if shared else "media",
        })
    return proposals


def _proposed_cargos(master: MasterList, skip: set[int]) -> list[dict]:
    """Rows whose `Organización` holds a job title.

    Rows already destined to disappear in a merge are skipped: correcting a
    cell and then deleting it is wasted, and the merge carries the original
    text into the survivor's notes anyway.
    """
    proposals = []
    for person in master.people:
        if person.row in skip or not person.organizacion:
            continue
        split = split_cargo(person.organizacion)
        if split.has_cargo:
            proposals.append({"person": person, "split": split})
    return proposals


def review(args: argparse.Namespace) -> int:
    master = MasterList.open(args.master)
    print(f"Master: {len(master.people)} personas "
          f"({'reestructurado' if master.is_restructured else 'formato antiguo'})")

    merges = _proposed_merges(master)
    cargos = _proposed_cargos(master, skip={m["drop"].row for m in merges})
    print(f"  {len(merges)} pares de filas que parecen la misma persona")
    print(f"  {len(cargos)} celdas de Organización que guardan un cargo")

    rows = [
        {
            "accion": "fusionar",
            "fila": m["keep"].row,
            "nombre": m["keep"].nombre,
            "actual": m["keep"].organizacion,
            "propuesta": m["keep"].organizacion,
            "cargo": "",
            "confianza": m["confianza"],
            "motivo": m["motivo"],
            "eliminar": m["drop"].row,
            "con": f"fila {m['drop'].row}: {m['drop'].nombre} / "
                   f"{m['drop'].organizacion or 'sin organización'} / "
                   f"{'; '.join(m['drop'].addresses) or 'sin email'}",
            "direcciones": "; ".join(dict.fromkeys(m["keep"].addresses + m["drop"].addresses)),
        }
        for m in merges
    ] + [
        {
            "accion": "cargo",
            "fila": c["person"].row,
            "nombre": c["person"].nombre,
            "actual": c["split"].original,
            "propuesta": c["split"].organizacion,
            "cargo": c["split"].cargo,
            "confianza": c["split"].confianza,
            "motivo": c["split"].motivo,
            "eliminar": "",
            "con": "",
            "direcciones": "; ".join(c["person"].addresses),
        }
        for c in cargos
    ]

    # Merges first, then the least certain splits: the sheet opens on whatever
    # most needs a decision, and a deletion always needs one.
    rows.sort(key=lambda r: (
        r["accion"] != "fusionar",
        CONFIDENCE_ORDER[r["confianza"]],
        r["nombre"].lower(),
    ))

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Curación"
    for index, name in enumerate(REVIEW_COLUMNS, start=1):
        cell = sheet.cell(1, index, name)
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS[name]

    for offset, row in enumerate(rows):
        values = [
            "SI", row["accion"], row["fila"], row["nombre"], row["actual"],
            row["propuesta"], row["cargo"], row["confianza"], row["motivo"],
            row["eliminar"], row["con"], row["direcciones"],
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(offset + 2, index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=index in (5, 9, 11, 12))
            if row["accion"] == "fusionar":
                cell.fill = FILL_FUSION
            elif row["confianza"] != "alta":
                cell.fill = FILL_DUDOSO

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_COLUMNS))}{len(rows) + 1}"
    output = Path(args.review_file)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)

    doubtful = sum(1 for r in rows if r["confianza"] != "alta")
    print(f"\n{doubtful} propuesta(s) con confianza media o baja — revísalas primero")
    print(f"Revisa y corrige: {output}")
    print("Luego: --apply")
    return 0


def _read_review(path: Path) -> Curation:
    sheet = openpyxl.load_workbook(path).active
    headers = [str(c.value or "").strip() for c in sheet[1]]
    index = {name: position for position, name in enumerate(headers)}
    missing = [c for c in ("Aplicar", "Acción", "Fila") if c not in index]
    if missing:
        raise SystemExit(f"{path.name}: faltan columnas {missing}. ¿Es el archivo de curación?")

    def value(row, name) -> str:
        position = index.get(name)
        if position is None or row[position].value is None:
            return ""
        text = str(row[position].value).strip()
        return "" if text.lower() == "nan" else text

    plan, skipped = Curation(), 0
    for row in sheet.iter_rows(min_row=2):
        if not any(c.value for c in row):
            continue
        if value(row, "Aplicar").upper() not in YES:
            skipped += 1
            continue
        target = int(float(value(row, "Fila")))
        organizacion = value(row, "Organización propuesta")
        cargo = value(row, "Cargo a Notas")
        if value(row, "Acción").lower() == "fusionar":
            plan.merges.append(Merge(
                keep=target,
                drop=int(float(value(row, "Fila a eliminar"))),
                organizacion=organizacion or None,
                notas=cargo or None,
            ))
        else:
            plan.amendments.append(Amendment(
                row=target,
                # An emptied cell is a decision too: "Curadora de arte" names
                # no organisation, so the column should say nothing.
                organizacion=organizacion,
                notas=cargo or None,
            ))
    print(f"Curación: {len(plan.amendments)} correcciones, {len(plan.merges)} fusiones, "
          f"{skipped} descartadas")
    return plan


def apply(args: argparse.Namespace) -> int:
    review_path = Path(args.review_file)
    if not review_path.exists():
        raise SystemExit(f"No existe {review_path}. Ejecuta primero --review.")
    plan = _read_review(review_path)
    if not (plan.amendments or plan.merges):
        print("Nada que curar.")
        return 0

    master = MasterList.open(args.master)
    if not master.is_restructured:
        restructured = master.restructure()
        print(f"Reestructurado a {len(master.headers)} columnas; "
              f"{len(restructured.split_addresses)} celdas con dos direcciones separadas; "
              f"{restructured.numbers_frozen} fórmulas de N convertidas a su valor")

    numbered = master.fill_missing_numbers()
    for row, number in numbered:
        print(f"  fila {row}: N asignado = {number}")

    left_alone = master.rows_without_address()
    if left_alone:
        print(f"{len(left_alone)} fila(s) sin dirección quedan intactas: "
              + ", ".join(f"fila {p.row} ({p.nombre})" for p in left_alone))

    before = master.last_data_row
    report = master.curate(plan)

    print(f"\n{len(report.organizaciones)} organización(es) corregida(s):")
    for nombre, was, now in report.organizaciones:
        print(f"  {nombre}: {was!r} -> {now or '(vacío)'!r}")
    print(f"{report.notas} nota(s) añadida(s)")
    for nombre, address in report.addresses_added:
        print(f"  {nombre}: dirección añadida {address}")
    for nombre, campo, value in report.carried_over:
        print(f"  {nombre}: {campo} conservado de la fila eliminada ({value})")
    for row, nombre in sorted(report.rows_removed):
        print(f"  fila {row} eliminada: {nombre}")
    print(f"\n{before} filas -> {master.last_data_row}")

    output = master.save_as(args.output or _default_output(args.master))
    print(f"Maestra curada -> {output}")
    print("\nEl archivo original no fue modificado.")
    return 0


def _default_output(master: Path) -> Path:
    return master.with_name(f"{master.stem}_curado{master.suffix}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Curar la lista maestra: cargos fuera de Organización y filas duplicadas.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--master", type=Path, required=True, help="lista maestra canónica")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--review", action="store_true", help="proponer (no escribe la maestra)")
    mode.add_argument("--apply", action="store_true", help="aplicar lo aprobado en la curación")
    parser.add_argument("--review-file", type=Path, default=Path("curacion.xlsx"),
                        help="hoja de curación (por defecto: curacion.xlsx)")
    parser.add_argument("--output", type=Path,
                        help="maestra curada (por defecto: *_curado.xlsx)")
    args = parser.parse_args()
    return review(args) if args.review else apply(args)


if __name__ == "__main__":
    raise SystemExit(main())
